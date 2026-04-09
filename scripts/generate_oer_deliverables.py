"""Generate OER test deliverables for class submission.

Creates two Excel files in deliverables/:
1) Model_Output_File.xlsx
2) Peer_Evaluation_File.xlsx
"""

from __future__ import annotations

import os
import sys
import subprocess
from datetime import date
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Ensure project root is importable when running this script directly.
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

from backend.oer_agent import OERAgent


CRITERIA = [
    "Open License",
    "Content Quality",
    "Accessibility",
    "Relevance to Course",
    "Currency/Up-to-date",
    "Pedagogical Value",
    "Technical Quality",
]


def _get_member_name(default: str = "mgligore") -> str:
    try:
        output = subprocess.check_output(["git", "config", "user.name"], text=True).strip()
        return output or default
    except Exception:
        return default


def collect_rows(min_rows: int = 10) -> List[Dict]:
    courses = ["ENGL 1101", "MATH 1111", "BIOL 1107", "HIST 2111"]
    agent = OERAgent(llm_provider="ollama", llm_model="qwen2.5:7b")

    rows: List[Dict] = []
    seen_urls = set()

    for course in courses:
        results = agent.find_oer_for_course(course)
        for item in results.get("evaluated_resources", []):
            resource = item.get("resource", {})
            url = (resource.get("url") or "").strip()
            if not url or url in seen_urls:
                continue

            rubric = item.get("rubric_evaluation", {})
            criterion_scores = rubric.get("criteria_evaluations", {})

            row = {
                "Course": course,
                "URL": url,
                "Title": resource.get("title", ""),
                "Overall Score": rubric.get("overall_score", ""),
            }

            for criterion in CRITERIA:
                row[criterion] = (criterion_scores.get(criterion, {}) or {}).get("score", "")

            rows.append(row)
            seen_urls.add(url)

            if len(rows) >= min_rows:
                return rows

    if len(rows) < min_rows:
        raise RuntimeError(f"Only collected {len(rows)} unique URLs; need at least {min_rows}.")

    return rows


def build_model_output(rows: List[Dict], team_name: str, member_name: str, run_date: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Model Output"

    metadata = [
        ["Team Name", team_name],
        ["Member Name(s)", member_name],
        ["Run Date", run_date],
        ["Model", "Ollama qwen2.5:7b"],
        ["Notes", "Agent-generated OER rubric scores (0-5)"],
    ]

    for row_idx, (label, value) in enumerate(metadata, start=1):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)

    header_row = 7
    headers = ["URL"] + CRITERIA + ["Overall Score", "Course", "Title"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start=header_row + 1):
        values = [row["URL"]] + [row[c] for c in CRITERIA] + [row["Overall Score"], row["Course"], row["Title"]]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1:
                cell.hyperlink = str(value)
                cell.style = "Hyperlink"
            if 2 <= col_idx <= 9:
                cell.alignment = Alignment(horizontal="center")

    widths = {
        1: 55,
        2: 14,
        3: 14,
        4: 14,
        5: 18,
        6: 18,
        7: 16,
        8: 14,
        9: 12,
        10: 12,
        11: 14,
        12: 42,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A8"

    output_path = os.path.join("deliverables", "Model_Output_File.xlsx")
    wb.save(output_path)
    return output_path


def build_peer_template(rows: List[Dict], team_name: str, member_name: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Peer Evaluation"

    metadata = [
        ["Team Name", team_name],
        ["Member Name(s)", member_name],
        ["Evaluator Name", "________________________"],
        ["Evaluation Date", "________________________"],
        [
            "Instructions",
            "Evaluate each URL using criteria definitions below. Do not view model scores. Use 0-5 where 5 is best.",
        ],
    ]

    for row_idx, (label, value) in enumerate(metadata, start=1):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)

    start_row = 8
    headers = ["URL", "Resource Title", "Course"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    for row_idx, row in enumerate(rows, start=start_row + 1):
        ws.cell(row=row_idx, column=1, value=row["URL"])
        ws.cell(row=row_idx, column=1).hyperlink = row["URL"]
        ws.cell(row=row_idx, column=1).style = "Hyperlink"
        ws.cell(row=row_idx, column=2, value=row["Title"])
        ws.cell(row=row_idx, column=3, value=row["Course"])

    guidance_row = start_row + len(rows) + 3
    ws.cell(row=guidance_row, column=1, value="Evaluation Criteria Guidance").font = Font(bold=True, size=12)

    guidance_headers = ["Category", "Maximum Score", "How to Evaluate (What to Look For)"]
    for col_idx, header in enumerate(guidance_headers, start=1):
        cell = ws.cell(row=guidance_row + 1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    guidance_rows = [
        (
            "Open License",
            "Max 5",
            "Check if resource clearly states an open license (e.g., CC BY, CC BY-SA, Public Domain). 5 = explicit open license; 3 = unclear; 1 = restrictive/all rights reserved.",
        ),
        (
            "Content Quality",
            "Max 5",
            "Review depth, accuracy, organization, and clarity. 5 = high-quality, well-structured, accurate content with useful examples.",
        ),
        (
            "Accessibility",
            "Max 5",
            "Check for accessible design: readable formatting, alt text/captions, screen-reader compatibility, downloadable formats. 5 = strong accessibility support.",
        ),
        (
            "Relevance to Course",
            "Max 5",
            "Compare against course topics and learning outcomes. 5 = directly supports major course units and objectives.",
        ),
        (
            "Currency/Up-to-date",
            "Max 5",
            "Look for publication/revision date and current references. 5 = recent/maintained; 3 = date unclear; 1 = obviously outdated.",
        ),
        (
            "Pedagogical Value",
            "Max 5",
            "Identify teaching utility: exercises, assessments, examples, learning activities, scaffolding. 5 = strong instructional usefulness.",
        ),
        (
            "Technical Quality",
            "Max 5",
            "Check link health, load speed, navigation, formatting, and media usability. 5 = stable, easy-to-use, minimal technical issues.",
        ),
    ]

    for row_idx, row in enumerate(guidance_rows, start=guidance_row + 2):
        ws.cell(row=row_idx, column=1, value=row[0])
        ws.cell(row=row_idx, column=2, value=row[1])
        ws.cell(row=row_idx, column=3, value=row[2])

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 90
    ws.freeze_panes = "A9"

    output_path = os.path.join("deliverables", "Peer_Evaluation_File.xlsx")
    wb.save(output_path)
    return output_path


def main() -> None:
    os.makedirs("deliverables", exist_ok=True)

    team_name = "Agentic OER Finder Team"
    member_name = _get_member_name()
    run_date = str(date.today())

    rows = collect_rows(min_rows=10)

    model_file = build_model_output(rows, team_name, member_name, run_date)
    peer_file = build_peer_template(rows, team_name, member_name)

    print(f"CREATED {model_file}")
    print(f"CREATED {peer_file}")
    print(f"ROWS {len(rows)}")
    for idx, row in enumerate(rows, start=1):
        print(f"{idx:02d}. {row['URL']} | overall={row['Overall Score']} | course={row['Course']}")


if __name__ == "__main__":
    main()
